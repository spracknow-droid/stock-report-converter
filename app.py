import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")

st.title("📊 원가수불부 자동 변환기")
st.markdown("파일을 업로드하면 **재고실사 양식**으로 변환됩니다.")

# --- 1. 사이드바: 파일 업로드 ---
st.sidebar.header("설정")
uploaded_file = st.sidebar.file_uploader("원가수불부 엑셀 파일을 업로드하세요", type=["xlsx", "xls"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    
    if len(df) > 0:
        df = df.drop(index=0).reset_index(drop=True)
    
    with st.spinner('데이터를 변환 중입니다...'):
        target_columns = ['품목계정그룹', '품목코드', '품목명', '단위', '기초재고', '입고계', '출고계', '기말재고']
        existing_cols = [col for col in target_columns if col in df.columns]
        transformed_df = df[existing_cols].copy()
        
        new_cols = ['실사수량 계', '저장위치1', '저장위치2', '저장위치3', '저장위치4']
        for col in new_cols:
            transformed_df[col] = "" 
            
        transformed_df[existing_cols] = transformed_df[existing_cols].fillna(0)
        
        rename_dict = {
            '기초재고': '11월말 수량',
            '입고계': '12월 입고 수량',
            '출고계': '12월 출고 수량',
            '기말재고': '기말 수량'
        }
        transformed_df.rename(columns=rename_dict, inplace=True)

    st.subheader("✅ 변환 완료 (미리보기)")
    st.dataframe(transformed_df.head(10), use_container_width=True)

    st.divider()
    
    output = BytesIO()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        groups = {
            '제품,상품': ['제품', '상품', '제품(OEM)'],
            '반제품': ['반제품'],
            '원재료': ['원재료'],
            '부재료': ['부재료']
        }

        num_format_cols = ['11월말 수량', '12월 입고 수량', '12월 출고 수량', '기말 수량']

        for sheet_name, filter_list in groups.items():
            if sheet_name == '제품,상품':
                sheet_df = transformed_df[transformed_df['품목계정그룹'].isin(filter_list)]
            else:
                sheet_df = transformed_df[transformed_df['품목계정그룹'] == filter_list[0]]

            if not sheet_df.empty:
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]
                
                for i, col_name in enumerate(sheet_df.columns):
                    column_letter = get_column_letter(i + 1)
                    
                    # [핵심 수정] 컬럼별 맞춤형 너비 설정
                    if col_name in ['품목계정그룹', '단위']:
                        adjusted_width = 12  # 짧은 텍스트 컬럼
                    elif col_name == '품목코드':
                        adjusted_width = 15  # 코드 컬럼
                    elif col_name == '품목명':
                        adjusted_width = 40  # 이름 컬럼 (가장 길게)
                    elif col_name in num_format_cols or col_name == '실사수량 계':
                        adjusted_width = 16  # 수량 및 실사 컬럼
                    else:
                        adjusted_width = 14  # 기타 위치 컬럼 등
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                    for row_idx in range(1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=i + 1)
                        cell.border = thin_border
                        
                        # 서식 및 정렬
                        if row_idx == 1:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            if col_name in num_format_cols:
                                cell.number_format = '#,##0'
                                cell.alignment = Alignment(horizontal='right')
                            else:
                                cell.alignment = Alignment(horizontal='center')

    st.download_button(
        label="📥 엑셀 다운로드",
        data=output.getvalue(),
        file_name="원가수불부_실사양식_최종.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("왼쪽 사이드바에서 원가수불부 엑셀 파일을 업로드해주세요.")
